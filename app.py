#!/usr/bin/env python3
"""
Researcher RCR Analyzer — Streamlit App
========================================
Upload an Excel file with researcher names and (optionally) institutions,
and this app queries PubMed + NIH iCite to produce per-researcher reports
of first/last-author papers with Relative Citation Ratio (RCR).

Run with:
    streamlit run app.py

Requirements:
    pip install streamlit openpyxl requests
"""

import io
import re
import time
import xml.etree.ElementTree as ET

import requests
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
ESEARCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
EFETCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
ICITE_URL = "https://icite.od.nih.gov/api/pubs"
NCBI_DELAY = 0.34
NCBI_DELAY_WITH_KEY = 0.11
ICITE_BATCH = 200


# ---------------------------------------------------------------------------
# Name helpers
# ---------------------------------------------------------------------------
def parse_name(raw_name):
    raw = raw_name.strip()
    if "," in raw:
        parts = [p.strip() for p in raw.split(",", 1)]
        last = parts[0]
        first_part = parts[1] if len(parts) > 1 else ""
    else:
        tokens = raw.split()
        if len(tokens) == 1:
            return {"last": tokens[0], "first": "", "initials": ""}
        if len(tokens[-1]) <= 3 and tokens[-1].isalpha() and tokens[-1].isupper() and len(tokens) == 2:
            return {"last": tokens[0], "first": "", "initials": tokens[-1]}
        last = tokens[-1]
        first_part = " ".join(tokens[:-1])

    first_tokens = first_part.split()
    initials = "".join(t[0].upper() for t in first_tokens if t)
    first = first_tokens[0] if first_tokens else ""
    return {"last": last, "first": first, "initials": initials}


def build_author_query(name_str, institutions):
    parsed = parse_name(name_str)
    last = parsed["last"]
    first = parsed["first"]
    initials = parsed["initials"]

    variants = set()
    if initials:
        variants.add(f"{last} {initials[0]}")
    if initials and len(initials) > 1:
        variants.add(f"{last} {initials}")
    if first:
        variants.add(f"{first} {last}")

    first_tokens = []
    raw = name_str.strip()
    if "," in raw:
        first_part = raw.split(",", 1)[1].strip()
        first_tokens = first_part.split()
    else:
        tokens = raw.split()
        if len(tokens) > 1:
            if len(tokens[-1]) <= 3 and tokens[-1].isalpha() and tokens[-1].isupper() and len(tokens) == 2:
                first_tokens = []
            else:
                first_tokens = tokens[:-1]

    if first_tokens and len(first_tokens) > 1:
        full_first = " ".join(first_tokens)
        variants.add(f"{full_first} {last}")
        mixed = first_tokens[0] + " " + " ".join(t[0] for t in first_tokens[1:])
        variants.add(f"{mixed} {last}")

    if first:
        variants.add(f"{last} {first}")
    if first_tokens and len(first_tokens) > 1:
        variants.add(f"{last} {' '.join(first_tokens)}")

    if not variants:
        variants.add(last)

    author_terms = [f'"{v}"' for v in sorted(variants)]
    author_query = "(" + " OR ".join(author_terms) + ")"

    if institutions:
        affil_terms = [f"{inst}[Affiliation]" for inst in institutions]
        affil_query = "(" + " OR ".join(affil_terms) + ")"
        return f"{author_query} AND {affil_query}"
    return author_query


def normalize_last(name_str):
    return parse_name(name_str)["last"].lower()


# ---------------------------------------------------------------------------
# PubMed
# ---------------------------------------------------------------------------
def search_pubmed(query, email, api_key=None):
    params = {"db": "pubmed", "term": query, "retmax": 5000, "retmode": "json", "email": email}
    if api_key:
        params["api_key"] = api_key
    resp = requests.get(ESEARCH_URL, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json().get("esearchresult", {}).get("idlist", [])


def fetch_article_details(pmids, email, api_key=None):
    articles = []
    delay = NCBI_DELAY_WITH_KEY if api_key else NCBI_DELAY
    for start in range(0, len(pmids), 200):
        batch = pmids[start:start + 200]
        params = {"db": "pubmed", "id": ",".join(batch), "retmode": "xml", "email": email}
        if api_key:
            params["api_key"] = api_key
        time.sleep(delay)
        resp = requests.get(EFETCH_URL, params=params, timeout=60)
        resp.raise_for_status()
        root = ET.fromstring(resp.content)
        for el in root.findall(".//PubmedArticle"):
            art = _parse_article_xml(el)
            if art:
                articles.append(art)
    return articles


def _parse_article_xml(article_el):
    medline = article_el.find("MedlineCitation")
    if medline is None:
        return None
    pmid_el = medline.find("PMID")
    pmid = pmid_el.text if pmid_el is not None else ""
    art = medline.find("Article")
    if art is None:
        return None

    title_el = art.find("ArticleTitle")
    title = "".join(title_el.itertext()) if title_el is not None else ""
    journal_el = art.find("Journal/Title")
    journal = journal_el.text if journal_el is not None else ""

    year = ""
    pub_date = art.find("Journal/JournalIssue/PubDate")
    if pub_date is not None:
        y_el = pub_date.find("Year")
        if y_el is not None:
            year = y_el.text
        else:
            md = pub_date.find("MedlineDate")
            if md is not None and md.text:
                m = re.match(r"(\d{4})", md.text)
                if m:
                    year = m.group(1)

    authors = []
    author_list = art.find("AuthorList")
    if author_list is not None:
        for au in author_list.findall("Author"):
            last_el = au.find("LastName")
            fore_el = au.find("ForeName")
            init_el = au.find("Initials")
            coll_el = au.find("CollectiveName")
            if last_el is not None and last_el.text:
                name = last_el.text
                if fore_el is not None and fore_el.text:
                    name = f"{last_el.text} {fore_el.text}"
                elif init_el is not None and init_el.text:
                    name = f"{last_el.text} {init_el.text}"
                authors.append({"full": name, "last": last_el.text.lower()})
            elif coll_el is not None and coll_el.text:
                authors.append({"full": coll_el.text, "last": coll_el.text.lower()})

    return {"pmid": pmid, "title": title, "journal": journal, "year": year, "authors": authors}


# ---------------------------------------------------------------------------
# First / last author filtering
# ---------------------------------------------------------------------------
def filter_first_last(articles, researcher_name):
    last_name = normalize_last(researcher_name)
    parsed = parse_name(researcher_name)
    initials = parsed["initials"].upper()
    first = parsed["first"].lower()

    results = []
    for art in articles:
        auths = art["authors"]
        if not auths:
            continue
        first_author = auths[0]
        last_author = auths[-1] if len(auths) > 1 else auths[0]
        role = _match_role(first_author, last_author, last_name, first, initials)
        if role:
            art_copy = dict(art)
            art_copy["author_position"] = role
            art_copy["author_list_str"] = "; ".join(a["full"] for a in auths)
            results.append(art_copy)
    return results


def _match_role(first_au, last_au, target_last, target_first, target_initials):
    is_first = _name_match(first_au, target_last, target_first, target_initials)
    is_last = _name_match(last_au, target_last, target_first, target_initials)
    if is_first and is_last:
        return "First & Last"
    if is_first:
        return "First"
    if is_last:
        return "Last"
    return None


def _name_match(author_dict, target_last, target_first, target_initials):
    if author_dict["last"] != target_last:
        return False
    if not target_first and not target_initials:
        return True
    full_lower = author_dict["full"].lower()
    if target_first and target_first in full_lower:
        return True
    if target_initials:
        parts = author_dict["full"].split()
        if len(parts) > 1:
            au_initials = "".join(p[0].upper() for p in parts[1:])
            if au_initials.startswith(target_initials):
                return True
    if not target_first and not target_initials:
        return True
    return False


# ---------------------------------------------------------------------------
# iCite
# ---------------------------------------------------------------------------
def fetch_icite(pmids):
    result = {}
    for start in range(0, len(pmids), ICITE_BATCH):
        batch = pmids[start:start + ICITE_BATCH]
        params = {
            "pmids": ",".join(batch),
            "fl": "pmid,relative_citation_ratio,nih_percentile,citation_count,"
                  "expected_citations_per_year,citations_per_year,is_research_article,"
                  "apt,is_clinical,provisional",
        }
        time.sleep(0.5)
        resp = requests.get(ICITE_URL, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        for pub in data.get("data", []):
            result[str(pub.get("pmid", ""))] = pub
    return result


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------
def compute_researcher_stats(articles, icite_data):
    rcrs = []
    years = []
    for art in articles:
        ic = icite_data.get(art["pmid"], {})
        rcr = ic.get("relative_citation_ratio")
        if rcr is not None:
            rcrs.append(rcr)
        yr = art.get("year", "")
        if yr and yr.isdigit():
            years.append(int(yr))

    rcr_sum = sum(rcrs) if rcrs else 0.0
    rcr_mean = (rcr_sum / len(rcrs)) if rcrs else 0.0

    if years and max(years) > min(years):
        year_span = max(years) - min(years)
    elif years:
        year_span = 1
    else:
        year_span = 1
    rcr_annual = rcr_sum / year_span

    return {
        "total_papers": len(articles),
        "rcr_sum": round(rcr_sum, 2),
        "rcr_mean": round(rcr_mean, 2),
        "rcr_annual": round(rcr_annual, 2),
        "year_first": min(years) if years else "N/A",
        "year_last": max(years) if years else "N/A",
    }


# ---------------------------------------------------------------------------
# Excel writers
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALT_FILL = PatternFill("solid", fgColor="D9E2F3")


def write_researcher_xlsx(researcher_name, institution_str, articles, icite_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "First & Last Author Papers"

    ws.merge_cells("A1:I1")
    ws["A1"] = f"{researcher_name} — {institution_str}" if institution_str else researcher_name
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = f"First & Last Author Papers with RCR — {len(articles)} papers found"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")

    headers = ["PMID", "Year", "Title", "Journal", "Authors", "Position", "RCR", "NIH Percentile", "Citation Count"]
    col_widths = [12, 8, 55, 25, 45, 14, 10, 16, 15]
    header_row = 4

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    def sort_key(a):
        ic = icite_data.get(a["pmid"], {})
        rcr = ic.get("relative_citation_ratio")
        yr = a.get("year", "")
        return (yr if yr else "0000", rcr if rcr is not None else -1)

    articles_sorted = sorted(articles, key=sort_key, reverse=True)

    for i, art in enumerate(articles_sorted):
        row = header_row + 1 + i
        pmid = art["pmid"]
        ic = icite_data.get(pmid, {})
        rcr = ic.get("relative_citation_ratio")
        percentile = ic.get("nih_percentile")
        citations = ic.get("citation_count")
        values = [
            int(pmid) if pmid.isdigit() else pmid,
            int(art["year"]) if art.get("year", "").isdigit() else art.get("year", ""),
            art["title"], art["journal"], art.get("author_list_str", ""),
            art.get("author_position", ""),
            round(rcr, 2) if rcr is not None else "N/A",
            round(percentile, 1) if percentile is not None else "N/A",
            citations if citations is not None else "N/A",
        ]
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = fill
            if col_idx in (1, 2, 6, 7, 8, 9):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if articles_sorted:
        sr = header_row + 1 + len(articles_sorted) + 1
        ws.cell(row=sr, column=5, value="Total papers:").font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=sr, column=6, value=len(articles_sorted)).font = Font(name="Arial", bold=True, size=10)
        valid_rcrs = [icite_data.get(a["pmid"], {}).get("relative_citation_ratio") for a in articles_sorted]
        valid_rcrs = [r for r in valid_rcrs if r is not None]
        if valid_rcrs:
            ws.cell(row=sr, column=7, value=round(sum(valid_rcrs) / len(valid_rcrs), 2)).font = Font(name="Arial", bold=True, size=10)
            ws.cell(row=sr + 1, column=5, value="Avg RCR (above)").font = Font(name="Arial", italic=True, size=9, color="666666")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:I{header_row + len(articles_sorted)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def write_summary_xlsx(summary_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "RCR Summary"

    ws.merge_cells("A1:H1")
    ws["A1"] = "Researcher RCR Summary"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Researcher", "Institution", "Papers", "RCR Sum", "Mean RCR", "Annual RCR", "First Pub Year", "Last Pub Year"]
    col_widths = [30, 30, 10, 12, 12, 14, 16, 16]
    header_row = 3

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for i, row_data in enumerate(summary_rows):
        row = header_row + 1 + i
        values = [
            row_data["name"], row_data["institution"], row_data["total_papers"],
            row_data["rcr_sum"], row_data["rcr_mean"], row_data["rcr_annual"],
            row_data["year_first"], row_data["year_last"],
        ]
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = fill
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{header_row}:H{header_row + len(summary_rows)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Input reader
# ---------------------------------------------------------------------------
def read_input_excel(uploaded_file):
    wb = load_workbook(uploaded_file, read_only=True)
    ws = wb.active
    researchers = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] is None:
            continue
        name = str(row[0]).strip()
        institutions = []
        for cell in row[1:]:
            if cell is not None:
                inst = str(cell).strip()
                if inst.lower() not in ("", "none", "n/a", "na"):
                    institutions.append(inst)
        if name.lower() in ("name", "researcher", "author"):
            continue
        researchers.append((name, institutions))
    wb.close()
    return researchers


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------
def run_pipeline(researchers, email, api_key, progress_bar, status_text):
    delay = NCBI_DELAY_WITH_KEY if api_key else NCBI_DELAY
    total = len(researchers)
    summary_rows = []
    individual_files = []

    for idx, (name, institutions) in enumerate(researchers):
        institution_str = "; ".join(institutions) if institutions else ""
        status_text.text(f"Processing {name}..." + (f" ({institution_str})" if institution_str else ""))

        # Search PubMed
        query = build_author_query(name, institutions)
        time.sleep(delay)
        try:
            pmids = search_pubmed(query, email, api_key=api_key)
        except Exception as e:
            st.warning(f"PubMed search failed for {name}: {e}")
            progress_bar.progress((idx + 1) / total)
            continue

        if not pmids:
            st.info(f"No articles found for {name} — skipping.")
            progress_bar.progress((idx + 1) / total)
            continue

        # Fetch details
        try:
            articles = fetch_article_details(pmids, email, api_key=api_key)
        except Exception as e:
            st.warning(f"Failed to fetch details for {name}: {e}")
            progress_bar.progress((idx + 1) / total)
            continue

        # Filter first/last
        first_last = filter_first_last(articles, name)
        if not first_last:
            st.info(f"No first/last author papers found for {name} — skipping.")
            progress_bar.progress((idx + 1) / total)
            continue

        # iCite
        fl_pmids = [a["pmid"] for a in first_last]
        try:
            icite_data = fetch_icite(fl_pmids)
        except Exception as e:
            st.warning(f"iCite query failed for {name}: {e}")
            icite_data = {}

        # Build individual xlsx
        xlsx_buf = write_researcher_xlsx(name, institution_str, first_last, icite_data)
        safe_name = re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')
        individual_files.append((f"{safe_name}_RCR_Report.xlsx", xlsx_buf))

        # Stats
        stats = compute_researcher_stats(first_last, icite_data)
        stats["name"] = name
        stats["institution"] = institution_str
        summary_rows.append(stats)

        progress_bar.progress((idx + 1) / total)

    return summary_rows, individual_files


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Researcher RCR Analyzer", page_icon="\U0001f4ca", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Source+Serif+4:wght@400;600;700&family=DM+Sans:wght@400;500;600&display=swap');
    .stApp { background-color: #f8f7f4; }
    .main-header { font-family: 'Source Serif 4', Georgia, serif; font-size: 2.4rem; font-weight: 700; color: #1a1a2e; margin-bottom: 0.2rem; letter-spacing: -0.02em; }
    .sub-header { font-family: 'DM Sans', sans-serif; font-size: 1.05rem; color: #6b7280; margin-bottom: 2rem; font-weight: 400; }
    .section-label { font-family: 'DM Sans', sans-serif; font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.08em; color: #9ca3af; margin-bottom: 0.5rem; }
    .stat-card { background: white; border: 1px solid #e5e7eb; border-radius: 12px; padding: 1.25rem 1.5rem; text-align: center; }
    .stat-value { font-family: 'Source Serif 4', Georgia, serif; font-size: 1.8rem; font-weight: 700; color: #1a1a2e; }
    .stat-label { font-family: 'DM Sans', sans-serif; font-size: 0.8rem; color: #9ca3af; text-transform: uppercase; letter-spacing: 0.05em; margin-top: 0.25rem; }
    div[data-testid="stDataFrame"] { border: 1px solid #e5e7eb; border-radius: 8px; }
    button[data-baseweb="tab"] { font-family: 'DM Sans', sans-serif !important; font-size: 1rem !important; font-weight: 500 !important; color: #6b7280 !important; padding: 0.75rem 1.25rem !important; }
    button[data-baseweb="tab"][aria-selected="true"] { color: #1a1a2e !important; font-weight: 600 !important; }
    button[data-baseweb="tab"]:hover { color: #1a1a2e !important; }
    div[data-baseweb="tab-list"] { gap: 0.5rem; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">Researcher RCR Analyzer</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Enter researcher names to retrieve their first &amp; last author publications and NIH Relative Citation Ratios.</div>', unsafe_allow_html=True)
st.markdown('<p style="font-family: DM Sans, sans-serif; font-size: 0.88rem; color: #9ca3af; margin-top: -1.2rem; margin-bottom: 0.3rem; line-height: 1.5;">This site will produce a summary with each researcher&rsquo;s total number of first- and last-author publications, the sum of their RCRs, their mean RCR, and their annual sum of RCRs.</p>', unsafe_allow_html=True)
st.markdown('<p style="font-family: DM Sans, sans-serif; font-size: 0.88rem; color: #9ca3af; margin-top: 0; margin-bottom: 2rem; line-height: 1.5;">It will also provide individual files listing papers and metrics for each researcher.</p>', unsafe_allow_html=True)

with st.sidebar:
    st.markdown('<div class="section-label">NCBI Credentials</div>', unsafe_allow_html=True)
    st.caption("Required by NCBI for E-utilities access. Your credentials are not stored.")
    email = st.text_input("Email address", placeholder="you@university.edu")
    api_key = st.text_input("NCBI API key (optional)", type="password", placeholder="Paste your key here")
    if api_key:
        st.success("API key provided — requests at 10/sec")
    else:
        st.info("No API key — requests limited to 3/sec")
    st.markdown("---")
    st.markdown('<div class="section-label">How to get an API key</div>', unsafe_allow_html=True)
    st.caption("1. Sign in at [ncbi.nlm.nih.gov](https://www.ncbi.nlm.nih.gov/myncbi/)\n2. Go to Account Settings\n3. Scroll to API Key Management\n4. Click Create API Key")

if "researchers" not in st.session_state:
    st.session_state.researchers = []

tab_manual, tab_upload = st.tabs(["Enter manually", "Upload Excel file"])

with tab_manual:
    st.markdown('<div class="section-label">ADD A RESEARCHER</div>', unsafe_allow_html=True)
    col_name, col_inst = st.columns([1, 2])
    with col_name:
        new_name = st.text_input("Researcher name", placeholder="e.g. Jane A Doe", key="input_name")
    with col_inst:
        new_institutions = st.text_input("Institution(s) — comma-separated, optional", placeholder="e.g. Utopia University, Springfield Medical Center", key="input_inst")

    if st.button("➕ Add researcher"):
        if new_name.strip():
            insts = [i.strip() for i in new_institutions.split(",") if i.strip()] if new_institutions else []
            st.session_state.researchers.append((new_name.strip(), insts))
            st.rerun()
        else:
            st.warning("Please enter a name.")

    st.markdown("")
    st.caption("**Tip:** Paste multiple researchers at once — one per line. Use a semicolon to separate name from institutions.")
    bulk_text = st.text_area("Bulk entry (optional)", placeholder="Jane A Doe; Utopia University, Springfield Medical Center\nJohn Q Smith; Greenfield Institute\nMaria Garcia", height=120, key="bulk_input")
    if st.button("➕ Add all from bulk entry"):
        added = 0
        for line in bulk_text.strip().splitlines():
            line = line.strip()
            if not line:
                continue
            if ";" in line:
                parts = line.split(";", 1)
                name = parts[0].strip()
                insts = [i.strip() for i in parts[1].split(",") if i.strip()]
            else:
                name = line
                insts = []
            if name:
                st.session_state.researchers.append((name, insts))
                added += 1
        if added:
            st.rerun()

with tab_upload:
    st.markdown('<div class="section-label">UPLOAD EXCEL FILE</div>', unsafe_allow_html=True)
    st.markdown(
        '<div style="font-size: 0.85rem; color: #6b7280; line-height: 1.7;">'
        'Upload an Excel file (.xlsx) formatted as follows:'
        '<ul style="margin-top: 0.3rem;">'
        '<li><strong>Column A</strong>: Researcher name (e.g. "Jane A Doe", "Doe, Jane A", or "Doe JA")</li>'
        '<li><strong>Columns B, C, D...</strong>: Institution(s), one per column (optional)</li>'
        '</ul>'
        'A header row (e.g. "Name", "Institution1", "Institution2") is auto-detected and skipped. '
        'Rows with a blank name are ignored. Institution cells that are blank, "N/A", or "None" are skipped.'
        '</div>',
        unsafe_allow_html=True,
    )
    uploaded_file = st.file_uploader("Choose file", type=["xlsx"], key="file_uploader")
    if uploaded_file:
        file_researchers = read_input_excel(uploaded_file)
        if file_researchers:
            if st.button(f"➕ Load {len(file_researchers)} researcher(s) from file"):
                st.session_state.researchers.extend(file_researchers)
                st.rerun()
        else:
            st.error("No researchers found in the file.")

researchers = st.session_state.researchers

if researchers:
    st.markdown("---")
    st.markdown(f'<div class="section-label">RESEARCHER LIST — {len(researchers)} TOTAL</div>', unsafe_allow_html=True)

    preview_data = []
    for i, (name, insts) in enumerate(researchers):
        preview_data.append({"#": i + 1, "Name": name, "Institution(s)": ", ".join(insts) if insts else "—"})
    st.dataframe(preview_data, use_container_width=True, hide_index=True)

    col_clear, col_run = st.columns([1, 3])
    with col_clear:
        if st.button("🗑️ Clear list", use_container_width=True):
            st.session_state.researchers = []
            st.rerun()
    with col_run:
        run_disabled = not email
        if not email:
            st.warning("Enter your email in the sidebar to run.")
        run_clicked = st.button("🚀 Run Analysis", type="primary", use_container_width=True, disabled=run_disabled)

    if run_clicked and email:
        st.markdown("---")
        progress_bar = st.progress(0)
        status_text = st.empty()
        summary_rows, individual_files = run_pipeline(researchers, email, api_key.strip() if api_key else None, progress_bar, status_text)
        status_text.text("Done!")

        if summary_rows:
            st.markdown("---")
            st.markdown('<div class="main-header" style="font-size:1.6rem;">Results</div>', unsafe_allow_html=True)
            total_papers = sum(r["total_papers"] for r in summary_rows)
            avg_rcr = sum(r["rcr_mean"] for r in summary_rows) / len(summary_rows) if summary_rows else 0
            total_rcr = sum(r["rcr_sum"] for r in summary_rows)

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.markdown(f'<div class="stat-card"><div class="stat-value">{len(summary_rows)}</div><div class="stat-label">Researchers</div></div>', unsafe_allow_html=True)
            with c2:
                st.markdown(f'<div class="stat-card"><div class="stat-value">{total_papers}</div><div class="stat-label">Total Papers</div></div>', unsafe_allow_html=True)
            with c3:
                st.markdown(f'<div class="stat-card"><div class="stat-value">{total_rcr:.1f}</div><div class="stat-label">Combined RCR</div></div>', unsafe_allow_html=True)
            with c4:
                st.markdown(f'<div class="stat-card"><div class="stat-value">{avg_rcr:.2f}</div><div class="stat-label">Avg Mean RCR</div></div>', unsafe_allow_html=True)

            st.markdown("")
            st.markdown('<div class="section-label">SUMMARY TABLE</div>', unsafe_allow_html=True)
            display_rows = []
            for r in sorted(summary_rows, key=lambda x: x["rcr_sum"], reverse=True):
                display_rows.append({"Researcher": r["name"], "Institution": r["institution"], "Papers": r["total_papers"], "RCR Sum": r["rcr_sum"], "Mean RCR": r["rcr_mean"], "Annual RCR": r["rcr_annual"], "First Year": r["year_first"], "Last Year": r["year_last"]})
            st.dataframe(display_rows, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown('<div class="section-label">DOWNLOADS</div>', unsafe_allow_html=True)
            summary_buf = write_summary_xlsx(summary_rows)
            st.download_button(label="📥 Download Summary Spreadsheet", data=summary_buf, file_name="RCR_Summary_All_Researchers.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            if individual_files:
                st.markdown("")
                st.markdown('<div class="section-label">INDIVIDUAL RESEARCHER REPORTS</div>', unsafe_allow_html=True)
                cols = st.columns(min(len(individual_files), 3))
                for i, (filename, buf) in enumerate(individual_files):
                    with cols[i % 3]:
                        display_name = filename.replace("_RCR_Report.xlsx", "").replace("_", " ")
                        st.download_button(label=f"📄 {display_name}", data=buf, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_{i}")
        else:
            st.warning("No results were generated. Check that the names and institutions are correct.")
